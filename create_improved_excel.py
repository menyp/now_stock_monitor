import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def create_improved_schedule():
    """Creates an improved version of the merge schedule Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Improved Merge Schedule"

    # --- 1. DATA --- #
    headers = [
        "Merge criteria", "354 (7/16-7/29)", "355 (7/30-8/12)", "356 (8/13-8/26)", 
        "357 (8/27-9/9)", "358 (9/10-9/23)", "359 (9/24-10/7)", "360 (10/8-10/21)", 
        "361 (10/22-11/4)", "362 (11/5-11/18)", "363 (11/19-12/2)"
    ]
    data = [
        ["Merge Slot", "", "12-Aug", "", "9-Sep", "", "7-Oct", "", "4-Nov", "", "24-Nov"],
        ["MVR clean", "Y", "N", "Y", "Y", "N", "Y", "Y", "Y", "N", "Y"],
        ["Client track test results", "<link>", "<link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>"],
        ["BE track test results", "<link>", "<link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>", "<Add a link>"],
        ["QE lead", "Alice", "Alice", "Bob", "Bob", "Charlie", "Charlie", "Alice", "Alice", "Bob", "Bob"],
        ["Dev lead", "Dave", "Dave", "Eve", "Eve", "Frank", "Frank", "Dave", "Dave", "Eve", "Eve"],
    ]

    # Add top-level title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "Dev Cycle (7/16 - 12/16)"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center')

    # Add headers
    ws.append(headers)

    # Add data rows
    for row in data:
        ws.append(row)

    # --- 2. UI/UX IMPROVEMENTS --- #

    # Freeze Panes at B3
    ws.freeze_panes = 'B3'

    # Alternating Row Colors (Banded Rows)
    light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for row in range(4, ws.max_row + 1, 2):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = light_gray_fill

    # Conditional Formatting for 'MVR clean' (Row 4)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add('B4:K4', Rule(type='cellIs', operator='equal', formula=['"Y"'], fill=green_fill))
    ws.conditional_formatting.add('B4:K4', Rule(type='cellIs', operator='equal', formula=['"N"'], fill=red_fill))

    # Highlight Current Dev Cycle Column (assuming today is around Aug 13)
    current_cycle_col = 'D'
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    for cell in ws[current_cycle_col]:
        cell.fill = yellow_fill

    # Data Validation for 'MVR clean' row
    dv = DataValidation(type="list", formula1='"Y,N,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('B4:K4')

    # --- 3. STYLING & FORMATTING --- #
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal='center')
    left_alignment = Alignment(horizontal='left')

    # Style header row (Row 2)
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Style first column (Merge criteria)
    first_col_font = Font(bold=True)
    for cell in ws['A']:
        cell.font = first_col_font
        cell.alignment = left_alignment

    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col_letter in [chr(c) for c in range(ord('B'), ord('L'))]:
        ws.column_dimensions[col_letter].width = 15

    # --- 4. SAVE FILE --- #
    output_filename = 'merge_schedule_improved.xlsx'
    wb.save(output_filename)
    print(f"Successfully created '{output_filename}' with UI/UX improvements.")

if __name__ == "__main__":
    create_improved_schedule()
